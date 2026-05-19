import os
import json
import tempfile
import fitz  # PyMuPDF
from flask import Flask, request, jsonify, send_file, render_template
from werkzeug.utils import secure_filename
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024  # 32MB
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'

GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY')
if not GEMINI_API_KEY:
    raise RuntimeError("GEMINI_API_KEY environment variable is not set")

genai.configure(api_key=GEMINI_API_KEY)

ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'webp', 'gif'}
IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp', 'gif'}
IMAGE_MIME_TYPES = {
    'png': 'image/png',
    'jpg': 'image/jpeg',
    'jpeg': 'image/jpeg',
    'webp': 'image/webp',
    'gif': 'image/gif',
}

SYSTEM_INSTRUCTIONS = """You are a data extraction expert. Your task is to analyze the provided content and extract all structured/tabular data from it.

Return ONLY a valid JSON object with this exact structure:
{
  "title": "descriptive title for the spreadsheet",
  "sheets": [
    {
      "name": "Sheet name",
      "headers": ["Column1", "Column2", "Column3"],
      "rows": [
        ["value1", "value2", "value3"],
        ["value1", "value2", "value3"]
      ]
    }
  ]
}

Rules:
- Extract ALL tables, lists, and structured data you find
- Each distinct table/section should be its own sheet
- Keep column headers concise and clear
- Preserve numeric values as numbers (no quotes)
- If there are multiple related tables, group them in the same sheet
- Make sure every row has the same number of values as headers
- Return ONLY the JSON, no markdown, no explanation
"""

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_extension(filename):
    return filename.rsplit('.', 1)[1].lower() if '.' in filename else ''

def extract_text_from_pdf(filepath):
    doc = fitz.open(filepath)
    text = ""
    for page in doc:
        text += page.get_text()
    doc.close()
    return text

def _parse_gemini_response(raw):
    raw = raw.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    return json.loads(raw.strip())

def call_gemini_text(content_text, user_prompt=""):
    model = genai.GenerativeModel('gemini-1.5-flash')
    prompt = f"{SYSTEM_INSTRUCTIONS}\n\nContent to extract:\n{content_text}"
    if user_prompt:
        prompt += f"\n\nAdditional instructions: {user_prompt}"
    response = model.generate_content(prompt)
    return _parse_gemini_response(response.text)

def call_gemini_image(image_bytes, mime_type, user_prompt=""):
    model = genai.GenerativeModel('gemini-2.5-flash')
    prompt = SYSTEM_INSTRUCTIONS
    if user_prompt:
        prompt += f"\n\nAdditional instructions: {user_prompt}"
    image_part = {"mime_type": mime_type, "data": image_bytes}
    response = model.generate_content([prompt, image_part])
    return _parse_gemini_response(response.text)

def build_excel(data: dict, output_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    HEADER_BG = "1A1A2E"
    HEADER_FG = "E8F4FD"
    ALT_ROW_BG = "F0F4FF"
    BORDER_COLOR = "C8D8F0"

    header_font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
    data_font = Font(name="Calibri", size=10, color="1A1A2E")
    title_font = Font(name="Calibri", bold=True, size=13, color=HEADER_BG)

    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR),
    )

    sheets_data = data.get("sheets", [])
    if not sheets_data:
        raise ValueError("No sheets data found")

    for sheet_info in sheets_data:
        ws = wb.create_sheet(title=sheet_info["name"][:31])
        headers = sheet_info.get("headers", [])
        rows = sheet_info.get("rows", [])

        ws.row_dimensions[1].height = 30
        title_cell = ws.cell(row=1, column=1, value=data.get("title", sheet_info["name"]))
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        if headers:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

        ws.row_dimensions[2].height = 22
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = PatternFill("solid", fgColor=HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx, row_data in enumerate(rows, start=3):
            ws.row_dimensions[row_idx].height = 18
            is_alt = (row_idx % 2 == 0)
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if is_alt:
                    cell.fill = PatternFill("solid", fgColor=ALT_ROW_BG)

        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(header))
            for row_data in rows:
                if col_idx - 1 < len(row_data):
                    max_len = max(max_len, len(str(row_data[col_idx - 1])))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    wb.save(output_path)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/convert', methods=['POST'])
def convert():
    user_prompt = request.form.get('prompt', '').strip()
    text_input = request.form.get('text_input', '').strip()

    uploaded_file = request.files.get('file')
    has_file = uploaded_file and uploaded_file.filename

    if not has_file and not text_input:
        return jsonify({'error': 'Please upload a file or paste some text'}), 400

    if has_file and not allowed_file(uploaded_file.filename):
        return jsonify({'error': 'Unsupported file type. Use PDF, PNG, JPG, WEBP, or GIF'}), 400

    try:
        data = None

        if has_file:
            ext = get_extension(uploaded_file.filename)
            filename = secure_filename(uploaded_file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            uploaded_file.save(filepath)

            try:
                if ext == 'pdf':
                    # PDF: extract text then send as text
                    pdf_text = extract_text_from_pdf(filepath)
                    combined = pdf_text
                    if text_input:
                        combined += f"\n\n--- Additional Text ---\n\n{text_input}"
                    data = call_gemini_text(combined, user_prompt)
                else:
                    # Image: send bytes directly to Gemini vision
                    with open(filepath, 'rb') as f:
                        image_bytes = f.read()
                    mime_type = IMAGE_MIME_TYPES[ext]
                    # If there's also pasted text, append it to the prompt
                    extra = f"\n\nAdditional text context:\n{text_input}" if text_input else ""
                    data = call_gemini_image(image_bytes, mime_type, user_prompt + extra)
            finally:
                os.remove(filepath)

        else:
            # Text only
            data = call_gemini_text(text_input, user_prompt)

    except json.JSONDecodeError:
        return jsonify({'error': 'Gemini returned invalid data. Try again or simplify the content.'}), 500
    except Exception as e:
        return jsonify({'error': f'Processing error: {str(e)}'}), 500

    output_filename = f"output_{os.getpid()}_{tempfile.gettempprefix()}.xlsx"
    output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)

    try:
        build_excel(data, output_path)
    except Exception as e:
        return jsonify({'error': f'Excel generation failed: {str(e)}'}), 500

    return send_file(
        output_path,
        as_attachment=True,
        download_name="extracted_data.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


if __name__ == '__main__':
    os.makedirs('uploads', exist_ok=True)
    os.makedirs('outputs', exist_ok=True)
    app.run(debug=True, port=5000)
